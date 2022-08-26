import face_recognition
import cv2
import numpy as np
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook("demo.xlsx")
#ws = wb.get_sheet_by_name(name = 'prog2') 

ws = wb.active
# import serial
# import time
# arduino = serial.Serial(port='COM5', baudrate=9600, timeout=.1)
# This is a tes of running face recognition on live video from your webcam. It's a little more complicated than the
# other example, but it includes some bARic performance tweaks to make things run a lot fARter:
#   1. Process each video frame at 1/4 resolution (though still display it at full resolution)
#   2. Only detect faces in every other frame of video.

# PLEARE NOTE: This example requires OpenCV (the `cv2` library) to be installed only to read from your webcam.
# OpenCV is *not* required to use the face_recognition library. It's only required if you want to run this
# specific tes. If you have trouble installing it, try any of the other tess that don't require it instead.

# Get a reference to webcam #0 (the default one)
video_capture = cv2.VideoCapture(0)

# Load a sample picture and learn how to recognize it.
fadlan_image = face_recognition.load_image_file("pas foto.jpg")
fadlan_face_encoding = face_recognition.face_encodings(fadlan_image)[0]

# Load a second sample picture and learn how to recognize it.
alya_image = face_recognition.load_image_file("alya.JPEG")
alya_face_encoding = face_recognition.face_encodings(alya_image)[0]
alya1_image = face_recognition.load_image_file("Alya Adibah.jpg")
alya1_face_encoding = face_recognition.face_encodings(alya1_image)[0]
sabrina_image = face_recognition.load_image_file("sabrina.JPEG")
sabrina_face_encoding = face_recognition.face_encodings(sabrina_image)[0]
sabrina1_image = face_recognition.load_image_file("sabrina1.jpg")
sabrina1_face_encoding = face_recognition.face_encodings(sabrina_image)[0]
el_image = face_recognition.load_image_file("el.JPEG")
el_face_encoding = face_recognition.face_encodings(el_image)[0]
awfa_image = face_recognition.load_image_file("awfa.JPEG")
awfa_face_encoding = face_recognition.face_encodings(awfa_image)[0]
awfa1_image = face_recognition.load_image_file("Kayla Awfa Nuha Wibowo.jpg")
awfa1_face_encoding = face_recognition.face_encodings(awfa1_image)[0]
zalfa_image = face_recognition.load_image_file("zalfa.JPEG")
zalfa_face_encoding = face_recognition.face_encodings(zalfa_image)[0]
zalfa1_image = face_recognition.load_image_file("zalfa2.jpeg")
zalfa1_face_encoding = face_recognition.face_encodings(zalfa_image)[0]
naura_image = face_recognition.load_image_file("naura.JPEG")
naura_face_encoding = face_recognition.face_encodings(naura_image)[0]
destina_image = face_recognition.load_image_file("destina.JPEG")
destina_face_encoding = face_recognition.face_encodings(destina_image)[0]
destina1_image = face_recognition.load_image_file("destina3.jpeg")
destina1_face_encoding = face_recognition.face_encodings(destina1_image)[0]
destina2_image = face_recognition.load_image_file("Destina Aqilah Husni.jpg")
destina2_face_encoding = face_recognition.face_encodings(destina2_image)[0]
qanita_image = face_recognition.load_image_file("qanita.JPEG")
qanita_face_encoding = face_recognition.face_encodings(qanita_image)[0]
azka_image = face_recognition.load_image_file("Dhiya Muhammad Azka.jpg")
azka_face_encoding = face_recognition.face_encodings(azka_image)[0]
azka1_image = face_recognition.load_image_file("azka3.jpeg")
azka1_face_encoding = face_recognition.face_encodings(azka1_image)[0]
azka2_image = face_recognition.load_image_file("Dhiya Muhammad Azka1.jpg")
azka2_face_encoding = face_recognition.face_encodings(azka2_image)[0]
kafi_image = face_recognition.load_image_file("Kafi Azhar kurniawan.jpg")
kafi_face_encoding = face_recognition.face_encodings(kafi_image)[0]
ijam_image = face_recognition.load_image_file("Azmi Abid Izzati.jpg")
ijam_face_encoding = face_recognition.face_encodings(ijam_image)[0]
sultan_image = face_recognition.load_image_file("Muhammad Sultan Wirakusuma.jpg")
sultan_face_encoding = face_recognition.face_encodings(sultan_image)[0]
muhammad_image = face_recognition.load_image_file("Muhammad Taufiq Rabbani.jpg")
muhammad_face_encoding = face_recognition.face_encodings(muhammad_image)[0]
ibnu_image = face_recognition.load_image_file("Ibnu Sina.jpg")
ibnu_face_encoding = face_recognition.face_encodings(ibnu_image)[0]
zahra_image = face_recognition.load_image_file("Zahra kaukaba.jpg")
zahra_face_encoding = face_recognition.face_encodings(zahra_image)[0]
zahra1_image = face_recognition.load_image_file("Zahra kaukaba1.jpg")
zahra1_face_encoding = face_recognition.face_encodings(zahra1_image)[0]
aisyah_image = face_recognition.load_image_file("Nur Aisyah Syarah.jpg")
aisyah_face_encoding = face_recognition.face_encodings(aisyah_image)[0]
caca_image = face_recognition.load_image_file("Salsabila Nur Hidayati.jpg")
caca_face_encoding = face_recognition.face_encodings(caca_image)[0]
kay_image = face_recognition.load_image_file("Kayyisah Althafunnisa.jpg")
kay_face_encoding = face_recognition.face_encodings(kay_image)[0]
kay1_image = face_recognition.load_image_file("Kayyisah Althafunnisa1.jpg")
kay1_face_encoding = face_recognition.face_encodings(kay_image)[0]
nuryana_image = face_recognition.load_image_file("Kayla Nuryana Hidayati.jpg")
nuryana_face_encoding = face_recognition.face_encodings(nuryana_image)[0]
nuryana1_image = face_recognition.load_image_file("Kayla Nuryana Hidayati1.jpg")
nuryana1_face_encoding = face_recognition.face_encodings(nuryana1_image)[0]
izul_image = face_recognition.load_image_file("Izzul.jpeg")
izul_face_encoding = face_recognition.face_encodings(izul_image)[0]

# Create arrays of known face encodings and their names
known_face_encodings = [
    fadlan_face_encoding,
    alya_face_encoding,
    alya1_face_encoding,
    sabrina_face_encoding,
    sabrina1_face_encoding,
    el_face_encoding,
    awfa_face_encoding,
    awfa1_face_encoding,
    zalfa_face_encoding,
    zalfa1_face_encoding,
    naura_face_encoding,
    destina_face_encoding,
    destina1_face_encoding,
    destina2_face_encoding,
    qanita_face_encoding,
    azka_face_encoding,
    azka1_face_encoding,
    azka2_face_encoding,
    kafi_face_encoding,
    ijam_face_encoding,
    sultan_face_encoding,
    muhammad_face_encoding,
    ibnu_face_encoding,
    zahra_face_encoding,
    zahra1_face_encoding,
    aisyah_face_encoding,
    caca_face_encoding,
    kay_face_encoding,
    kay1_face_encoding,
    nuryana_face_encoding,
    nuryana1_face_encoding,
    izul_face_encoding

]
known_face_names = [
    "Fadhlan",
    "Alya",
    "Alya",
    "Sabrina",
    "Sabrina1",
    "El Syauqi",
    "Kayla Awfa",
    "Kayla Awfa",
    "Zalfa",
    "Zalfa",
    "Naura",
    "Destina",
    "Destina",
    "Destina",
    "Qanita",
    "Azka",
    "Azka",
    "Azka",
    "Kafi",
    "Izam",
    "Sultan",
    "Muhammad",
    "Ibnu",
    "Zahra",
    "Zahra",
    "Aisyah",
    "Salsabila",
    "Kayyisah",
    "Kayyisah",
    "Kayla Nuryana",
    "Kayla Nuryana",
    "Izzul"
]

# Initialize some variables
face_locations = []
face_encodings = []
face_names = []
process_this_frame = True

while True:
    # Grab a single frame of video
    ret, frame = video_capture.read(1)

    # Resize frame of video to 1/4 size for fARter face recognition processing
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

    # Convert the image from ARR color (which OpenCV uses) to RGB color (which face_recognition uses)
    rgb_small_frame = small_frame[:, :, ::-1]

    # Only process every other frame of video to save time
    if process_this_frame:
        # Find all the faces and face encodings in the current frame of video
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

        face_names = []
        for face_encoding in face_encodings:
            # See if the face is a match for the known face(s)
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Unknown"



            # # If a match wAR found in known_face_encodings, just use the first one.
            # if True in matches:
            #     first_match_index = matches.index(True)
            #     name = known_face_names[first_match_index]

            # Or instead, use the known face with the smallest distance to the new face
            face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
            best_match_index = np.argmin(face_distances)
            if matches[best_match_index]:
                name = known_face_names[best_match_index]            
            face_names.append(name)
            x = datetime.datetime.now()
            if x.strftime("%x") == datetime.datetime(2022, 3, 26).strftime("%x"):
                z='BM'
            if x.strftime("%x") == datetime.datetime(2022, 3, 28).strftime("%x"):
                z='BN'
                y=int(x.strftime("%H"))
                if y>17:
                   z='BO' 
            if x.strftime("%x") == datetime.datetime(2022, 3, 29).strftime("%x"):
                z='BP'
                y=int(x.strftime("%H"))
                if y>17:
                   z='BQ' 
            if x.strftime("%x") == datetime.datetime(2022, 3, 30).strftime("%x"):
                z='BR'
                y=int(x.strftime("%H"))
                if y>17:
                   z='BS' 
            if x.strftime("%x") == datetime.datetime(2022, 3, 14).strftime("%x"):
                z='BK'
            if x.strftime("%x") == datetime.datetime(2022, 3, 15).strftime("%x"):
                z='BB'
                y=int(x.strftime("%H"))
                if y>17:
                   z='BC'
            if x.strftime("%x") == datetime.datetime(2022, 3, 16).strftime("%x"):
                z='BD'
                y=int(x.strftime("%H"))
                if y>17:
                   z='BE'  
            if name == "Destina":
                ws[z+'15'] = "H"
                cv2.imwrite('destina2.jpg', frame)
                wb.save("demo.xlsx")
            if name == "Zalfa":
                ws[z+'17'] = "H"
                cv2.imwrite('Zalfa2.jpg', frame)
                wb.save("demo.xlsx")
            if name == "Qanita":
                ws[z+'22'] = "H"
                cv2.imwrite('qanita2.jpg', frame)
                wb.save("demo.xlsx")
            if name == "Alya":
                ws[z+'20'] = "H"
                cv2.imwrite('alya2.jpg', frame)
                wb.save("demo.xlsx")
            if name == "Kayla Awfa":
                ws[z+'18'] = "H"
                cv2.imwrite('awfa2.jpg', frame)
                wb.save("demo.xlsx")
            if name == "Sabrina":
                ws[z+'21'] = "H"
                cv2.imwrite('sabrina2.jpg', frame)
                wb.save("demo.xlsx")
            if name == "El Syauqi":
                ws[z+'5'] = "H"
                cv2.imwrite('el2.jpg', frame)
            if name == "Azka":
                ws[z+'4'] = "H"
                cv2.imwrite('azka2.jpg', frame)
            if name == "Naura":
                ws[z+'19'] = "H"
                cv2.imwrite('naura2.jpg', frame)
            if name == "Kafi":
                ws[z+'6'] = "H"
                cv2.imwrite('kafi2.jpg', frame)
            if name == "Izam":
                ws[z+'7'] = "H"
                cv2.imwrite('izam2.jpg', frame)
            if name == "Sultan":
                ws[z+'8'] = "H"
                cv2.imwrite('sultan2.jpg', frame)
            if name == "Muhammad":
                ws[z+'10'] = "H"
                cv2.imwrite('muhammad2.jpg', frame)
            if name == "Ibnu":
                ws[z+'9'] = "H"
                cv2.imwrite('ibnu2.jpg', frame)
            if name == "Zahra":
                ws[z+'14'] = "H"
                cv2.imwrite('zahra2.jpg', frame)
            if name == "Aisyah":
                ws[z+'13'] = "H"
                cv2.imwrite('aisyah2.jpg', frame)
            if name == "Salsabila":
                ws[z+'23'] = "H"
                cv2.imwrite('salsabila2.jpg', frame)
            if name == "Kayyisah":
                ws[z+'12'] = "H"
                cv2.imwrite('kayyisah2.jpg', frame)
            if name == "Kayla Nuryana":
                ws[z+'11'] = "H"
                cv2.imwrite('nuryana2.jpg', frame)
            if name == "Izzul":
                ws[z+'24'] = "H"
                cv2.imwrite('Izzul2.jpg', frame)

            wb.save("demo.xlsx")
           
    process_this_frame = not process_this_frame


    # Display the results
    for (top, right, bottom, left), name in zip(face_locations, face_names):
        # Scale back up face locations since the frame we detected in wAR scaled to 1/4 size
        top *= 4
        right *= 4
        bottom *= 4
        left *= 4

        # Draw a box around the face
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

        # Draw a label with a name below the face
        cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
        font = cv2.FONT_HERSHEY_DUPLEX
        cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)

    # Display the resulting image
    cv2.imshow('Video', frame)

    # Hit 'q' on the keyboard to quit!
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# ReleARe handle to the webcam
video_capture.release()
cv2.destroyAllWindows()